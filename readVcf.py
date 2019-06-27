import io
from os import listdir, getcwd
from os.path import isfile, join
import openpyxl
from tkinter import *
from tkinter import messagebox

dirpath = getcwd()
exelCellsContentTypes = ['', 'ORG', 'ADR', 'TITLE', '', 'N', 'FN', '', '', 'TEL', 'EMAIL']

def getDataFromVcf(path):
    with io.open(path, encoding='utf-8') as file:
        contact = {}
        for line in file:
            if "PHOTO" in line:
                break
            key = line.split(":")[0] # Line Format:  N;CHARSET=utf-8:Иванюк;Василий;Валентинович;;
            key = key.split(";")[0]
            value = line.split(":")[1]
            contact[key] = value   
    print(contact.get("N"))
    writeToExel(contact)

def getVcfFilesPathes():
    vcfFiles = [f for f in listdir(dirpath) if isfile(join(dirpath, f)) and f.endswith(".vcf") ]
    return vcfFiles

def getXlsXFilePath():
    return [f for f in listdir(dirpath) if isfile(join(dirpath, f)) and f.endswith(".xlsx") ][0]

def processVcfs():
    vcfFiles = getVcfFilesPathes()
    for vcf in vcfFiles:
        getDataFromVcf(vcf)

def writeToExel(data):
    file = getXlsXFilePath()
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.active #todo unhardcode
    rowContent = list(exelCellsContentTypes)
    for entry in data:
        for i, val in enumerate(exelCellsContentTypes):
            if entry == val:
                if entry == "ADR":
                    data[entry] = data[entry].split(';')[-4] #get city
                if entry == "N":
                    data[entry] = data[entry].split(';')[0] #get surname
                rowContent[i]=(data[entry])
                print(rowContent)
                break
    print(rowContent)
    max = ws.max_row
    for i, entry in enumerate(rowContent,start=1):
        print(i, entry)
        ws.cell(row=max+1, column=i, value=entry)
    wb.save(file)

def main():
    processVcfs()
    root.quit()





def quit():
    root.quit()
 
root = Tk()
root.title("Программа для копирования контактов из визиток в таблицу")
root.geometry("600x200")

text = Label(width=500, height=200, justify=LEFT, font="Arial 12" )
text['text']="Положите в папку визитки в формате vcf\n и таблицу exel и нажимите ОК.\n\nТаблица должна быть закрыта."
text.pack()
 
message_button = Button(text="OK", command=main, font="Arial 12")
message_button.place(relx=.8, rely=.8, anchor="c")

message_button = Button(text="Отмена", command=quit, font="Arial 12")
message_button.place(relx=.9, rely=.8, anchor="c")
 
root.mainloop()